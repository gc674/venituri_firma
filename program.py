# Exercitiul II
# O firma a realizat in ultimii ani urmatoarele cifre de afaceri:
# 2010 → 230000 EURO
# 2011 → 280000 EURO
# 2012 → 310000 EURO
#
# 2013 → 320000 EURO
# 2014 → 350000 EURO
# 2015 → 310000 EURO
# 2016 → 380000 EURO
# 2017 → 400000 EURO
# a) Folosind modulul openpyxl creati un fisier de tip Excel si salvati in acesta cifrele
# de afacere ale firmei.
# b) Adaugati intr-o celula formula care calculeaza suma cifrelor de afacere ale firmei
# pt. anii 2010-2017.
# c) Folositi un alt stil pt. celula care contine suma (la alegere alta culoare, alt font,
# boldat etc)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


cifra_afaceri = [(2010, 230000), (2011, 280000), (2012, 310000), (2013, 320000), (2014, 350000),
                 (2015, 310000), (2016, 380000), (2017, 400000)]

align = Alignment(horizontal='center', vertical='bottom')

wb = openpyxl.Workbook()
ws = wb.active

ws.title = "Cifra afaceri 2010-1017"

# print(wb.sheetnames)

# cap tabel
header = ['An', 'Cifra de afaceri', 'Total cifra afaceri']

for i in range(1, len(header)+1):
    cell = ws.cell(row=1, column=i, value=header[i-1])
    cell.fill = PatternFill('solid', fgColor="00C0C0C0")
    if i == 1:
        ws.column_dimensions[get_column_letter(i)].width = 8
    else:
        ws.column_dimensions[get_column_letter(i)].width = len(header[i-1])

# se scrie formula sum total în C9
total = ws['C9']
total.value = f'=sum(B2:B{len(cifra_afaceri)})'
total.font = Font(bold=True)
total.fill = PatternFill('solid', fgColor="0099CCFF")
total.alignment = align

# se scriu datele în tabel incrementând liniile
l = 2

for i, j in cifra_afaceri:
    ws.cell(row=l, column=1, value=i).alignment = align
    ws.cell(row=l, column=2, value=j).alignment = align
    l += 1

# se citesc datele din tabel memorie
# for row in ws.iter_rows(min_row=1, max_row=len(cifra_afaceri), max_col=2, values_only=True):
#     print(row)

wb.save('firma.xlsx')